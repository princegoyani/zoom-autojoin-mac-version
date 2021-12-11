import time
from typing import ValuesView
import openpyxl as xl 
from datetime import datetime 
from inpt import studclas , submb , subippe

def daycheck(day):#row
    clasaddrow = 0
    print("Today is " , day)
    for row in range (4,sheet.max_row+1):
        daycell = sheet.cell(row,1)
        d = daycell.value
        #print("d = " , d)
        if day == d :
            clasaddrow = row
            #print("row" ,"=" , clasaddrow)
            #k=3
            row = row +1
            tryday =  sheet.cell(row,1)
            for h in range(1,4):
                tryday =  sheet.cell(row,1)
                #print("td = " , tryday.value)
                if tryday.value != None :
                    diff = (row+h)-row
                    #print("Diff = ", diff)
                    break
                row = row +1 
            break
        
    return clasaddrow , diff
                   
def checktime(hour,min):#colume
    clasaddcol = 0

    print("Time Right now " , hour ," h" ,min ,"min")
    
    #ckmin = clasmin.value
    for i in range(2, sheet.max_column,4):
        #print(i)
        if hour == sheet.cell(2,i).value:
            print("hour match",i,"a")
            if hour == sheet.cell(2,i+2).value:
                print("hour match",i,"b")
                if min >= sheet.cell(2,i+3).value:
                    i = i + 4
                    print("hour match",i,"c")
                    continue
                elif min <= sheet.cell(2,i+1).value:
                    if min < sheet.cell(2,i-1).value and hour == sheet.cell(2,i-2).value:
                        print("hour match",i,"d")   
                        clasaddcol = i - 3
                        print("hour match",i,"d")   
                        break
                    else:
                        clasaddcol= i + 1
                        print("hour match",i,"e")                       
                        break
                        
                else:
                    clasaddcol= i + 1
                    break
            else:
                clasaddcol= i + 1
                break
        elif hour == sheet.cell(2,i+2).value:
                print("hour match",i,"f")
                if min <= sheet.cell(2,i+3).value:
                        clasaddcol= i + 1
                        print("hour match",i,"g")                       
                        break
                else:
                    continue



    #print("i= " , i, "sheet.cell(2,i).value = " , sheet.cell(2,i).value, "now = " , nowhour)
    #else:
     #   print("Came Late !!")
      #  raise TimeoutError
    print("col = " , clasaddcol)

    return clasaddcol

def checkclass(row,col,stmb,stippe):
    clasadd = 0
    zoomid = 0
    col = clasaddcol
    row = clasaddrow
    #print("row = " , row, "col = " , col)
    l = 0
    sub = {}
    for i in range(row,(row+diff)):
        #print(i)
        clasadd = sheet.cell(i,col)
        #print("Classadd" , clasadd.value) 
        if clasadd.value != None:
            #print("Class is there" , clasadd.value)
            l = l+1
            a = i
            sub[sheet.cell(i , col-1).value.upper()] = i
            # print("l = " , l)
           
        """    elif i == (row+diff)+1:
                print("NO Class NOw")
                break
    

                #print(sub)
    else:
        print("NO Class Right Now !!")"""
    
    

    if l != 1 and  l != 0:
        #print("Select Ur Subject : " )

        for p in sub.keys() : 
            #print(p)
            if p == stmb:
                studsub = stmb
                break
            elif p == stippe:
                studsub = stippe
                break
        else:
            #print("Enter Ur subject correctly !!!!")
            time.sleep(3)
            raise ValuesView
        

        if studsub in sub.keys():
            print(studsub)
            row = sub[studsub]
            zoomid = sheet.cell(row,col).value
            if zoomid == 0 :
                raise ValueError
            print("zoom = ", zoomid)
        else:
            print("Wrong Entered !!")
    elif l == 0 :
        print("No Class Right now !! ")
    else:
        print("Class add row=",a,"col = " , col)
        zoomid = sheet.cell(a,col).value
        print("zoom = ", zoomid)
    
    return clasadd , zoomid

""" # password is fixed so no need of this function
def password():

    global teachid
    global teachpass
    for j in range(2,passheet.max_row+1):
        teachid = passheet.cell(j,1).value
        if zoomid == teachid : 
            #print(teachid)
            teachpass = passheet.cell(j,2).value
            #print(teachpass)
            break
    else:
        print("Password not exist!!")
         """   

wb = xl.load_workbook("timetable.xlsx")
sheet = wb[studclas.value]
clashour = sheet["b2"]
#clasmin =  sheet["c2"]
daycell = sheet.cell(4,1)

c = datetime.now()
tday = c.strftime("%A")
nowhour = int(datetime.now().hour)
nowmin = int(datetime.now().minute)


""" # password duciotn
passbook = xl.load_workbook("teachers id.xlsx")
passheet = passbook["Sheet1"]
"""
"""
clasadd=0
clasaddcol=0
clasaddrow =0
zoomid=0
diff=0
teachid = 0 
teachpass = 0

"""

row = daycheck(tday)
print(row)                             #row
clasaddrow = row[0]                              #row address
diff = row[1]                                    #if two subject are there in a cell ex. maths / bio 
clasaddcol = checktime(nowhour,nowmin)           #column
classaddtemp = checkclass(clasaddrow , clasaddcol, submb.value , subippe.value)
clasadd = classaddtemp[0]
zoomid = classaddtemp[1]

#password()



#password()
#way( zoomid , teachpass)      

#print(clastime.value)   

